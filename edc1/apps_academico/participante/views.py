from django.shortcuts import render
from django.views.generic import TemplateView
from rest_framework.response import Response
from rest_framework.decorators import *
from rest_framework.renderers import *
from django.http import JsonResponse
from .models import *
from .serializers import *
from django.core.exceptions import ObjectDoesNotExist
import datetime
from django.views.generic import View
#-------import para reportes----
from django.conf import settings
from io import BytesIO
from django.shortcuts import render, HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.colors import HexColor, yellow, black
from reportlab.lib.styles import getSampleStyleSheet
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle, TA_CENTER
from reportlab.lib.enums import TA_JUSTIFY, TA_LEFT, TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.utils import ImageReader
from reportlab.lib.units import inch, mm, cm
from reportlab.platypus import (
    Paragraph,
    Table,
    SimpleDocTemplate,
    Spacer,
    TableStyle,
    Paragraph)
from openpyxl import Workbook
from openpyxl.styles import Alignment,Border,Font,PatternFill,Side,Color
from apps_academico.docente.models import *

from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.template import Context
from apps_academico.reporte_aca.utils import link_callback
#----------------------------------
# Create your views here.

@api_view(['GET'])
@renderer_classes([JSONRenderer])
def asistencia_by_evento_and_fecha(request):
    evento_id = request.GET.get('evento_id', None)
    fecha = request.GET.get('fecha', None)

    try:
        fecha_object = datetime.datetime.strptime(fecha, "%Y-%m-%d").date()
        asistencia = Asistencia.objects.get(
            evento_id=evento_id,
            fecha=fecha_object
        )
        serializer = AsistenciaSerializer(asistencia)
        return Response(serializer.data)
    except ValueError:
        return JsonResponse({
            'detail': 'Either evento or fecha is null'
        }, status=400)
    except TypeError:
        return JsonResponse({
            'detail': 'Fecha parameter was missing'
        }, status=400)
    except ObjectDoesNotExist:
        return JsonResponse({
            'detail': '. Not found'
        }, status=404)

#---------Seccion de reportes-------------------------


class ParticipantesReprobados(View):
    style = getSampleStyleSheet()

    def cabecera(self, pdf):
        width, height = A4
        archivo_imagen = settings.MEDIA_ROOT+'/image_event/espol.png'
        pdf.drawImage(archivo_imagen, 10, 760, 230,
                      90, preserveAspectRatio=True)
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(410, 810, b"Reporte de eventos por genero")
        pdf.setFillColor(yellow)
        pdf.rect(520, 791, 67, 12, fill=True, stroke=False)
        pdf.setFillColor(black)
        pdf.setFont("Helvetica", 10)
        pdf.drawString(520, 793, u"POEC0502 V7")

    def pie_pagina(self, pdf):
        pdf.setFont("Helvetica", 10)
        now = datetime.now()
        pdf.drawString(
            10, 45, 'CEC ESPOL, Campus Gustavo Galindo Velasco | Teléf:042269763 | 0960507588 ')
        pdf.drawString(10, 30, u"Fecha impresión:"+str(now.day) +
                       '/'+str(now.month)+'/'+str(now.year))
        page_num = pdf.getPageNumber()
        text = "Pág. %s|1" % page_num
        pdf.drawString(500, 30, text)
        pdf.drawString(200, 30, u'Usuario: ')
        pdf.drawString(240, 30, u'Luis Eduardo Ardila Macias')
        pdf.setFillColor(HexColor('#3c5634'))
        pdf.drawString(10, 60, "//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////")

    def contenido(self, pdf, y):
        width, height = A4
        styles = getSampleStyleSheet()
        styleN = styles["BodyText"]
        styleN.alignment = TA_CENTER
        Numero = Paragraph('<b>N.-</b>', styleN)
        Codigo = Paragraph('<b>Codigo</b>', styleN)
        N_evento = Paragraph('<b>Nombre del evento</b>', styleN)
        Cedula = Paragraph('<b>Cedula</b>', styleN)
        N_part = Paragraph('<b>Nombre del Participante</b>',styleN)
        Asistencia = Paragraph('<b>Asistencia</b>', styleN)
        N_final = Paragraph('<b>Nota final</b>', styleN)
        encabezado1 = [
            [Numero, Codigo, N_evento, Cedula, N_part , Asistencia, N_final],
            ['1', '', '', '','', '%', ''],
            ['2', '', '', '','',  '', ''],
            ['3', '', '', '','',  '', ''],
            ['4', '', '', '','',  '', '']
        ]

        t = Table(encabezado1, colWidths=[
                  1*cm, 2*cm, 5*cm, 3.1*cm,5*cm, 2.3*cm,2*cm])
        t.setStyle(TableStyle([
            ('BOX', (0, 0), (-1, -1), 0.20, colors.black),
            ('INNERGRID', (0, 0), (-1, -1), 0.10, colors.black),
            ('BOTTOMPADDING', (0, 0), (5,0), 10),
        ]))

        t.wrapOn(pdf, width, height)
        t.drawOn(pdf, 8, 650)
        #Formulario Inferior
        styleB = styles["BodyText"]
        styleB.alignment = TA_LEFT
        var_1 = Paragraph('<b>Total reprobados por asistencia</b>', styleB)
        var_2 = Paragraph('<b>Total por Evento</b>', styleB)
        var_3 = Paragraph('<b>Total reprobados por  calificación</b>', styleB)
        var_4 = Paragraph('<b>Total inscritos personalmente</b>',styleB)
        var_5 = Paragraph('<b>Total inscritos Auspiciados </b>', styleB)
        form_inferior = [
                          [var_1 ,'Σ total','Asistencia minina del 80%',var_2,'Σ total'],
                          [var_3,'Σ total','Nota mínima 70/100',var_4,'Σ total'],
                          ['', '', '', var_5, 'Σ total']
                        ]
        t_form = Table(form_inferior, colWidths=[
                  5*cm, 2*cm, 4.5*cm, 5*cm,4*cm])
        t_form.setStyle(TableStyle([
            ('BOTTOMPADDING', (3, 0), (3,0),15),
            ('BOTTOMPADDING', (1, 0), (1,0),15),
            ('BOTTOMPADDING', (1, 1), (1,1),15),
            ('BOTTOMPADDING', (2, 0), (2,0),15),
            ('BOTTOMPADDING', (2, 1), (2,1), 15),
            ('BOTTOMPADDING', (4, 0), (4, 0), 15),
            ('BOTTOMPADDING', (4, 1), (4, 1), 15),
            ('BOTTOMPADDING', (4, 2), (4, 2), 15),
        ]))

        t_form.wrapOn(pdf, width, height)
        t_form.drawOn(pdf, 8, 90)


    def get(self, request, ):
        response = HttpResponse(content_type='application/pdf')
        buffer = BytesIO()
        pdf = canvas.Canvas(buffer, pagesize=A4)
        y = 590
        self.cabecera(pdf)
        self.pie_pagina(pdf)
        self.contenido(pdf, y)
        pdf.showPage()
        pdf.save()
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response

#---------------------------------------
class Perfil_participante(View):
    def cabecera(self,pdf):
        archivo_imagen = settings.MEDIA_ROOT+'/image_event/espol.png'
        pdf.drawImage(archivo_imagen, 40, 740, 190, 90,preserveAspectRatio=True)
        pdf.line(260,740,35,740)
        pdf.setFont("Times-Roman", 10)
        pdf.drawString(390, 790, b" PERFIL DE PARTICIPANTES")
        pdf.drawString(426, 774, u"CÓDIGO EVENTO ")
        pdf.drawString(466, 761, u"########")
        pdf.drawString(35, 720, u"Programa:") ; pdf.drawString(260, 720, u"Duración:")
        pdf.drawString(35, 705, u"Promoción:") ; pdf.drawString(260, 705, u"Fecha Inicio:")
        pdf.drawString(35, 690, u"Curso:") ; pdf.drawString(260, 690, u"Fecha Final:")
        pdf.drawString(35, 675, u"Instructor:") ; pdf.drawString(260, 675, u"Tipo de Capacitación:")


    def pie_pagina(self,pdf):
        pdf.setFillColor(HexColor(308011))
        pdf.drawString(10, 60, u" /////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////")
        pdf.setFillColor(HexColor(000000))
        page_num = pdf.getPageNumber()
        text = "Pag %s |1 " % page_num
        pdf.drawString(500, 30, text)
        pdf.drawString(35,50 ,u'CEC ESPOL, Campus Gustavo Galindo Velasco | Teléf:042269763 | 0960507588 ')
        now = datetime.now()
        pdf.drawString(35,35, u"Fecha impresión:"+str(now.day)+'/'+str(now.month)+'/'+str(now.year))
        pdf.drawString(260, 35,u'Usuario')

    def tabla(self,pdf,y):
        styles = getSampleStyleSheet()
        styleN = styles["BodyText"]
        styleN.alignment = TA_LEFT
        styleBH = styles["Normal"]
        styleBH.alignment = TA_CENTER
        hnum = Paragraph('''<b>N°.-</b>''', styleBH)
        hcedula = Paragraph('''<b>Cedula</b>''', styleBH)
        hnombre = Paragraph('''<b>Nombre</b>''', styleBH)
        htitulo3 = Paragraph('''<b>Titulo tercer nivel</b>''', styleBH)
        htitulo4 = Paragraph('''<b>Titulo cuarto nivel</b>''', styleBH)
        hprofesion = Paragraph('''<b>Profesion</b>''', styleBH)
        hcargo = Paragraph('''<b>Cargo</b>''', styleBH)
        harea = Paragraph('''<b>Área</b>''', styleBH)
        encabezados = (hnum,hcedula,hnombre,htitulo3,htitulo4,hprofesion,hcargo,harea)
        detalles = [('0')]
        detalle_orden = Table([encabezados] + detalles, colWidths=[0.9 * cm, 2 * cm, 2 * cm, 2.5 * cm,2.5*cm,2.5*cm,2*cm])
        detalle_orden.setStyle(TableStyle(
            [
                ('ALIGN',(0,0),(2,0),'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black), 
                #('SPAN',(-1,-1),(-3-1))
                ('FONTSIZE', (0, 0), (-1, -1), 10),
            ]
        ))
        detalle_orden.wrapOn(pdf, 850, 650)
        detalle_orden.drawOn(pdf, 60,y)
    def get(self, request, ):
            response = HttpResponse(content_type='application/pdf')
            buffer = BytesIO()
            pdf = canvas.Canvas(buffer)
            y = 600
            self.cabecera(pdf)
            self.pie_pagina(pdf)
            self.tabla(pdf,y)
            pdf.showPage()
            pdf.save()
            pdf = buffer.getvalue()
            buffer.close()
            response.write(pdf)
            return response


#--------------------------------
class Conctacto_participante(View):
    def cabecera(self,pdf):
        archivo_imagen = settings.MEDIA_ROOT+'/image_event/espol.png'
        pdf.drawImage(archivo_imagen, 40, 740, 190, 90,preserveAspectRatio=True)
        pdf.line(260,740,35,740)
        pdf.setFont("Times-Roman", 10)
        pdf.drawString(365, 790, b"CONTACTO DE PARTICIPANTE")
        pdf.drawString(426, 774, u"CÓDIGO EVENTO ")
        pdf.drawString(466, 761, u"########")
        pdf.drawString(35, 720, u"Evento:") ; pdf.drawString(260, 720, u"Aula:"); pdf.drawString(410, 720, u"Horario:")
        pdf.drawString(35, 705, u"Promoción:") ; pdf.drawString(260, 705, u"Fecha Inicio:")
        pdf.drawString(35, 690, u"Módulo:") ; pdf.drawString(260, 690, u"Fecha Final:")
        pdf.drawString(35, 675, u"Tipo de capacitación:") ; pdf.drawString(260, 675, u"Duración:")


    def pie_pagina(self,pdf):
        pdf.setFillColor(HexColor(308011))
        pdf.drawString(10, 60, u" /////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////")
        pdf.setFillColor(HexColor(000000))
        page_num = pdf.getPageNumber()
        text = "Pag %s" % page_num
        pdf.drawString(500, 30, text)
        pdf.drawString(35,50 ,u'CEC ESPOL, Campus Gustavo Galindo Velasco | Teléf:042269763 | 0960507588 ')
        now = datetime.now()
        pdf.drawString(35,35, u"Fecha impresión:"+str(now.day)+'/'+str(now.month)+'/'+str(now.year))
        pdf.drawString(260, 35,u'Usuario')

    def tabla(self,pdf,y):
        width, height = A4
        styles = getSampleStyleSheet()
        styleN = styles["BodyText"]
        styleN.alignment = TA_LEFT
        styleBH = styles["Normal"]
        styleBH.alignment = TA_CENTER
        hparticipante = Paragraph('''<b>PARTICIPANTE</b>''', styleBH)
        hemail1 = Paragraph('''<b>Email1</b>''', styleBH)
        hemail2 = Paragraph('''<b>Email2</b>''', styleBH)
        hcelular = Paragraph('''<b>Celular</b>''', styleBH)
        htelf_domicilio = Paragraph('''<b>Telefono Domicilio</b>''', styleBH)
        htelf_trabajo = Paragraph('''<b>Telefono Trabajo</b>''', styleBH)
        hdireccion = Paragraph('''<b>Dirección</b>''', styleBH)
        nombre= Paragraph('''Steen''',styleBH)
        #harea = Paragraph('''<b>Área</b>''', styleBH)
        encabezados = ('',hparticipante,hemail1,hemail2,hcelular,htelf_domicilio,htelf_trabajo,hdireccion)
        detalles = [('1',nombre,'')]
        #Establecemos el tamaño de cada una de las columnas de la tabla
        detalle_orden = Table([encabezados] + detalles, colWidths=[0.5*cm,3.5 * cm, 2 * cm, 2 * cm, 2.5 * cm,2.5*cm,2.5*cm,2.5*cm])
        #Aplicamos estilos a las celdas de la tabla
        detalle_orden.setStyle(TableStyle(
            [
                #La primera fila(encabezados) va a estar centrada
                ('ALIGN',(0,0),(2,0),'CENTER'),
                #Los bordes de todas las celdas serán de color negro y con un grosor de 1
                ('BOX', (1, 0), (-1, -1), 0.20, colors.black),
                ('GRID', (1, 0), (-1,-1), 0.5, colors.black), 
                 ('GRID', (0, 1), (0,1), 0.5, colors.black), 
                #El tamaño de las letras de cada una de las celdas será de 10
                ('FONTSIZE', (0, 0), (-1, -1), 10),
            ]
        ))
        #Establecemos el tamaño de la hoja que ocupará la tabla 
        detalle_orden.wrapOn(pdf,width, height)
        
        #Definimos la coordenada donde se dibujará la tabla
        detalle_orden.drawOn(pdf,60,y)
    def get(self, request, ):
            
            #Indicamos el tipo de contenido a devolver, en este caso un pdf
            response = HttpResponse(content_type='application/pdf')
            #La clase io.BytesIO permite tratar un array de bytes como un fichero binario, se utiliza como almacenamiento temporal
            buffer = BytesIO()
            #Canvas nos permite hacer el reporte con coordenadas X y Y
            pdf = canvas.Canvas(buffer)
            #Llamo al método donde están definidos los datos que aparecen en el reporte.
            y = 590 
            self.cabecera(pdf)
            self.pie_pagina(pdf)
            self.tabla(pdf,y)
            #Con show page hacemos un corte de página para pasar a la siguiente
            pdf.showPage()
            pdf.save()
            pdf = buffer.getvalue()
            buffer.close()
            response.write(pdf)
            return response

def contacto_participante(request):
    template_path = 'reporte/contacto_participante.html'
    #design = DesignEvento.objects.filter()
    #context = {'design':design}
    response = HttpResponse(content_type='application/pdf')
    #response['Content-Disposition'] = 'attachment; filename="report.pdf"'
    template = get_template(template_path)
    html = template.render()
    pisaStatus = pisa.CreatePDF(
       html, dest=response, link_callback=link_callback)
    if pisaStatus.err:
       return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response

def historico_participante(usuarios):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Hoja' + str()
    # ---------------------------------para darle diseño a mi titulo en la hoja-----------------------------------------


    # ---------------------------------------cambiar caracteristicas de las celdas--------------------------------------
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 20
    ws.column_dimensions['L'].width = 20
    ws.column_dimensions['M'].width = 20
    ws.column_dimensions['N'].width = 20
    ws.column_dimensions['O'].width = 20
    ws.column_dimensions['P'].width = 20
    ws.column_dimensions['Q'].width = 20
    ws.column_dimensions['R'].width = 20
    ws.column_dimensions['S'].width = 20
    ws.column_dimensions['U'].width = 20
    ws.column_dimensions['V'].width = 20
    ws.column_dimensions['W'].width = 20
    ws.column_dimensions['X'].width = 20
    ws.column_dimensions['Y'].width = 20
    ws.column_dimensions['Z'].width = 20
    ws.column_dimensions['AA'].width = 20
    ws.column_dimensions['AB'].width = 25
    #ws.column_dimensions['D'].width = 20
    # ----------------------------------------------------darle diseño a mi cabecera------------------------------------
    columnas = (
                      'A3',
                      'B3',
                      'C3',
                      'D3',
                      'E3',
                      'F3',
                      'G3',
                      'H3',
                      'I3',
                      'J3',
                      'K3',
                      'L3',
                      'M3',
                      'N3',
                      'O3',
                      'P3',
                      'Q3',
                      'R3',
                      'S3',
                      'T3',
                      'U3',
                      'V3',
                      'W3',
                      'X3',
                      'Y3',
                      'Z3',
                      'AA3',
                    )
    valores_columnas = (
                      'Secuencia',
                      'Curso',
                      'Tipo Capacitacion',
                      'Evento programa',
                      'Promoción',
                      'Fecha inicio',
                      'Fecha fin',
                      'Horario',
                      'Estado',
                      'Costo',
                      'Total horas',
                      'Instructor',
                      'Área',
                      'Tipo certificado',
                      'Asistencia',
                      'Nota',
                      'Certificado recibido',
                      'Modalidad',
                      'Codigo diseño',
                      'Versión',
                      'Tipo inscripción',
                      'Empresa',
                      'Aula',
                      'Asesor responsable',
                      'vendedor',
                      'Coordinador responsable',
                      'Fecha creacion evento',
                      'Fecha registro',  
                    )

    limite_header = ''
    limite = ''

    data = {
            'header':columnas,
            'data':valores_columnas
           }

    
    for col,val in zip(columnas,valores_columnas):
        ws[col].alignment = Alignment(horizontal="center", vertical="center")
        ws[col].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws[col].fill = PatternFill(start_color='608652', end_color='608652', fill_type="solid")
        ws[col].font = Font(name='Calibri', size=12, bold=True)
        ws[col] = val
        limite = col

    pos_col = 0
    
    for indice in range(len(limite)):
        if indice == 3:
            limite = '{}{}1'.format(limite[0],limite[1])
        else:
            limite = '{}1'.format(limite[0])

    ws.merge_cells('A1:{}'.format(limite))
    ws.row_dimensions[1].height = 25
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))

    ws['A1'].fill = PatternFill(start_color='608652', end_color='608652', fill_type="solid")
    ws['A1'].font = Font(name='Calibri', size=12, bold=True)
    ws['A1'] = 'Historial del participante'



    """
    for head in valores_header:
        ws[head].alignment = Alignment(horizontal="center", vertical="center")
        ws[head].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws[head].fill = PatternFill(start_color='608652', end_color='608652', fill_type="solid")
        ws[head].font = Font(name='Calibri', size=12, bold=True)
        ws[head] = 'Secuencia'

    """
    """
    ws['A3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['A3'].fill = PatternFill(start_color='608652', end_color='608652', fill_type="solid")
    ws['A3'].font = Font(name='Calibri', size=12, bold=True)
    ws['A3'] = 'Secuencia'

    ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['B3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['B3'].fill = PatternFill(start_color='608652', end_color='608652', fill_type="solid")
    ws['B3'].font = Font(name='Calibri', size=12, bold=True)
    ws['B3'] = 'Curso'

    ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['C3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['C3'].fill = PatternFill(start_color='608652', end_color='608652', fill_type="solid")
    ws['C3'].font = Font(name='Calibri', size=12, bold=True)
    ws['C3'] = 'Tipo de capacitacion'

    ws['D3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['D3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                            top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['D3'].fill = PatternFill(start_color='608652', end_color='608652', fill_type="solid")
    ws['D3'].font = Font(name='Calibri', size=12, bold=True)
    ws['D3'] = 'Evento programa'

    ws['E3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['E3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                            top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['E3'].fill = PatternFill(start_color='608652', end_color='608652', fill_type="solid")
    ws['E3'].font = Font(name='Calibri', size=12, bold=True)
    ws['E3'] = 'Promoción'

    ws['F3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['F3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                            top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['F3'].fill = PatternFill(start_color='608652', end_color='608652', fill_type="solid")
    ws['F3'].font = Font(name='Calibri', size=12, bold=True)
    ws['F3'] = 'Fecha de inicio'

    ws['G3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['G3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                            top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['G3'].fill = PatternFill(start_color='608652', end_color='608652', fill_type="solid")
    ws['G3'].font = Font(name='Calibri', size=12, bold=True)
    ws['G3'] = 'Fecha fin'

    ws['H3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['H3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                            top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['H3'].fill = PatternFill(start_color='608652', end_color='608652', fill_type="solid")
    ws['H3'].font = Font(name='Calibri', size=12, bold=True)
    ws['H3'] = 'Horario'

    ws['I3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['I3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                            top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['I3'].fill = PatternFill(start_color='608652', end_color='608652', fill_type="solid")
    ws['I3'].font = Font(name='Calibri', size=12, bold=True)
    ws['I3'] = 'Horario'

    ws['J3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['J3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                            top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['J3'].fill = PatternFill(start_color='608652', end_color='608652', fill_type="solid")
    ws['J3'].font = Font(name='Calibri', size=12, bold=True)
    ws['J3'] = 'Estado'

    ws['K3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['K3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                            top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['K3'].fill = PatternFill(start_color='608652', end_color='608652', fill_type="solid")
    ws['K3'].font = Font(name='Calibri', size=12, bold=True)
    ws['K3'] = 'Costo'

    ws['L3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['L3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                           top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['L3'].fill = PatternFill(start_color='608652', end_color='608652', fill_type="solid")
    ws['L3'].font = Font(name='Calibri', size=12, bold=True)
    ws['L3'] = 'Total horas'

    ws['M3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['M3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                          top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['M3'].fill = PatternFill(start_color='608652', end_color='608652', fill_type="solid")
    ws['M3'].font = Font(name='Calibri', size=12, bold=True)
    ws['M3'] = 'Instructor'

    ws['N3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['N3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                           top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['N3'].fill = PatternFill(start_color='608652', end_color='608652', fill_type="solid")
    ws['N3'].font = Font(name='Calibri', size=12, bold=True)
    ws['N3'] = 'Total horas'

    ws['O3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['O3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                           top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['O3'].fill = PatternFill(start_color='608652', end_color='608652', fill_type="solid")
    ws['O3'].font = Font(name='Calibri', size=12, bold=True)
    ws['O3'] = 'Instructor'

    ws['P3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['P3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                            top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['P3'].fill = PatternFill(start_color='608652', end_color='608652', fill_type="solid")
    ws['P3'].font = Font(name='Calibri', size=12, bold=True)
    ws['P3'] = 'Área'

    ws['Q3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['Q3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                            top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['Q3'].fill = PatternFill(start_color='608652', end_color='608652', fill_type="solid")
    ws['Q3'].font = Font(name='Calibri', size=12, bold=True)
    ws['Q3'] = 'Tipo certificado'

    ws['R3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['R3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                            top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['R3'].fill = PatternFill(start_color='608652', end_color='608652', fill_type="solid")
    ws['R3'].font = Font(name='Calibri', size=12, bold=True)
    ws['R3'] = 'Asistencia'

    ws['S3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['S3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                            top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['S3'].fill = PatternFill(start_color='608652', end_color='608652', fill_type="solid")
    ws['S3'].font = Font(name='Calibri', size=12, bold=True)
    ws['S3'] = 'Nota'

    ws['T3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['T3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                            top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['T3'].fill = PatternFill(start_color='608652', end_color='608652', fill_type="solid")
    ws['T3'].font = Font(name='Calibri', size=12, bold=True)
    ws['T3'] = 'Certificado recibido'

    ws['U3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['U3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                           top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['U3'].fill = PatternFill(start_color='608652', end_color='608652', fill_type="solid")
    ws['U3'].font = Font(name='Calibri', size=12, bold=True)
    ws['U3'] = 'Certificado recibido'

    ws['U3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['U3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                           top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['U3'].fill = PatternFill(start_color='608652', end_color='608652', fill_type="solid")
    ws['U3'].font = Font(name='Calibri', size=12, bold=True)
    ws['U3'] = 'Certificado recibido'
    """

    # ---------------------------pintar datos en excel y EXPORTAR DATOS DE LA BD---------------------------------------------------------
    
    """
    controlador = 4
    for confusuario in usuarios:
        ws.cell(row=controlador, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=1).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=1).font = Font(name='Calibri', size=8)
        ws.cell(row=controlador, column=1).value = confusuario.usuario

        ws.cell(row=controlador, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=2).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=2).font = Font(name='Calibri', size=8)
        ws.cell(row=controlador,column=2).value = confusuario.id_persona.nombres + " " + confusuario.id_persona.apellidos

        ws.cell(row=controlador, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=3).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=3).font = Font(name='Calibri', size=8)
        ws.cell(row=controlador, column=3).value = confusuario.id_genr_tipo_usuario.nombre
        controlador += 1
    # cont += 1
    """

    # establecer el nombre de mi archivo
    nombre_archivo = "ReportePersonalizadoExcel.xlsx"
    # Definir tipo de respuesta que va a dar
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename = {0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response


def part_reprobados(request):
    template_path = 'reportes/reprobados_por_evento.html'
    #design = DesignEvento.objects.filter()
    context = {'design':'#'}
    response = HttpResponse(content_type='application/pdf')
    #response['Content-Disposition'] = 'attachment; filename="report.pdf"'
    template = get_template(template_path)
    html = template.render(context)
    pisaStatus = pisa.CreatePDF(
       html, dest=response, link_callback=link_callback)
    if pisaStatus.err:
       return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response